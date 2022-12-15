import requests
import re
import openpyxl
import time
#配置headers
headers={
        'cookie': 'dom3ic8zudi28v8lr6fgphwffqoz0j6c=eff981d2-4b16-470c-8657-6c24b290b988:3:1; localized=1; _ga=GA1.2.7467935.1670460630; _ga_Z3V6T9VBM6=GS1.1.1670512561.5.1.1670516155.0.0.0; __cf_bm=UtKX8sOkgOMJinKc8VXqsvqA5IcYcPiqbsNUPAkPaTw-1670628610-0-AT0f7fWU8xSd4YhWjCTO7HtRM/fWFVtYWvUmM/AcWi64rnENHRCxngKhXJYDGqAsAU4PBHo9sNCpXn2CvJ3r5UrMiVFMTm+5sLIjO5AKcjGsDh88FpQ69rZOt542TeVY+CtYEPUGMIGMm4lFONC7DK4=; search_history=["%E6%9D%BE%E6%9C%AC%E3%81%84%E3%81%A1%E3%81%8B","DVDES-644","%E4%BD%93%E6%A3%80","%E6%80%A7%E6%95%99%E8%82%B2","MTALL-031"]; XSRF-TOKEN=eyJpdiI6IisrZU5CUHlSOW8xaEhRUHRSdUhIMnc9PSIsInZhbHVlIjoiQTNVM1FYbWRkRWlaaWlibTNRclFOZGx5UDVyallENml5VlYxTy9ubmF5QlJLcHBBZFM2YkZ5VWQzWVFaYUZhaW94dDJkSVpIL2Rvd3FzbU9aS2YzODZyRHVJSTNrVytBWFd0TFVSV1NzZlNTME05YWFoSnBNRDcyalZva1dLaFgiLCJtYWMiOiIyOWJlZTI5NDBiYTkxZjk4OTU1M2U3YzAyNjg0ZThiZmFjMmZkMzc3YTY0NzU3ZmEyYTY3MDQ2MGEwZGE5ZGI4IiwidGFnIjoiIn0=; missav_session=eyJpdiI6IjBuWDlteG5vTlhmS2ppYk5oM0pSckE9PSIsInZhbHVlIjoiVGt4VjliYVJoc1RZRWZOQVBRSitnd09JOWFyVkY5TzBJMEt6RWVVeUs4cjlxcnkreFArUUNuUDBOMVFseXkwNDNzR0hEREN0emJyb2VSbWtvL3FIWjVmQWlDalJtRTM2QmhrTnNnenppQmQ5RDAxaHBtWkVGcWQ0SkNEZXhXUGoiLCJtYWMiOiJlYWQ0MGU2MGRmMjE4NWY2MGJiNDVjZDkzMDViY2JjYTYzZGI2OTIxOTI1OGQ2NDY5YjAwMmNmNzBhNThiM2NmIiwidGFnIjoiIn0=; CdHKodX58fXKDhwjz5xS1bqwjfrwa4oRgSDiOizw=eyJpdiI6IjMvSXlyb3RtcklSeUtxSEhnR0g5SGc9PSIsInZhbHVlIjoiQ3U5M3Q0UDVsWnFmQXRodnNqVTVNQW5rUGZwV3NCTlJheWJlS00rNXBIZnI5YzZmTFBTMUNRcWd1S1puaTk5a1RkckhXd0tNQi9IWmJJc3RWUVRGcjJFL1cvT0tLd0draExTTXlYZHdWa2dHaGw5dEJLTXI5WHFtNWVDTGErWlhUMllPN0FEL1Q2Q29kVUh1WHdlMzFxZ1dwVmcrSzlGSXA0aVB3ZGxzK0JoU09XTHU1N0ZJNXVRbHFHeWxSQTM0SlFMTWthTDBhampwdTVNUWp6YmttQ1p3cDFzdzRnVWY4UjVFRktJU2JOZ3BsK2RrZzJUayt2ZG5XZHF1cUVNbEd1UjBleHovamhqU3pid0VyQ1ZnUmFHYlJDNmxyS2t6V0xsQzVkY1dpTDRDSVg3MlArWXo4Qk4xWk5jQUpNQTQvWjdlTDBPMHhxSE1iMGxIeS9aTW1zRUlyU2RxVldaS0RqaFNUekM2TVpmMXM3Ry9IMHI3RjNsVHJwRU9QWm5mcGs5TzlKYkMrbkc1cTNkb3JacFNtZz09IiwibWFjIjoiZWRmMDVkMmQxYWJlNWE3ZDY4ZGYyMTk0MGYzN2YwMzUyMTQwNDFmNTI5YTY1MTM1OTc4ZWQwODdhMzg2ZDMyMSIsInRhZyI6IiJ9',
        'referer': 'https://missav.com',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
}
#主页面URL配置
url='https://missav.com/actresses/ranking'
#解析主网页
r=requests.get(url)
html=r.text
#提取数据（女优名字）
__name__=re.findall('<h4 class="text-nord13 truncate">(.*?)</h4>',html)


#-----------------------------以下为excel对接--------------------------------
# 创建一个Excel workbook 对象
book = openpyxl.Workbook()
#---------------------------------对接完成-----------------------------------



#主要部分   main
for i in range(len(__name__)):
    #进入每个女优对应的主页，并解析数据
    name=__name__[i]
    url=f'https://missav.com/actresses/{name}?sort=views&page=1'
    print(url)
    r=requests.get(url,headers=headers)
    html=r.text

    #提取数据提取出视频名称和视频网页链接
    title_url=re.findall('<a class="text-secondary group-hover:text-primary" href="(.*?)">.*?</a>',html,re.S)
    title=re.findall('<a class="text-secondary group-hover:text-primary" href=".*?">(.*?)</a>',html,re.S)

    #创建表单（以女优名字命名）
    sh = book.create_sheet(f'{name}',i)
    # 写标题栏
    sh['A1'] =  '片名'
    sh['B1'] =  '视频网页地址'
    sh['C1'] =  '视频PT下载链接'
    row = 2

    #开始写入数据
    for a in range(len(title)):
        html_get=requests.get(title_url[a],headers=headers)
        html_get=html_get.text
        time.sleep(0.5)
        #解析每个视频播放链接网页内容
        try:    #尝试寻找pt下载链接
            result=re.search('<a href="https://mypikpak.com/drive/.*?;__add_url=(.*?)" target=".*?" rel=".*?" class=".*?">',html_get,re.S,verify=False)
            url_pt=result.group(1)
        except Exception:
            url_pt='很抱歉，无法正确解析视频链接'
        sh.cell(row, 1).value = title[a]        #写入视频标题
        sh.cell(row, 2).value = title_url[a]    #写入视频播放链接
        sh.cell(row, 3).value = url_pt          #写入视频pt下载链接
        row += 1


# Excel保存文件
book.save('AV女优大全.xlsx')
